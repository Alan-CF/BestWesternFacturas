import zipfile
import streamlit as st
from io import BytesIO
import pandas as pd
import openpyxl
import tempfile
import os
from typing import List

class FileProcessor:
    def __init__(self):
        self.start_row = 17
        self.template_path = 'Template Amazon.xlsx'

        self.facturasWB = openpyxl.load_workbook('Template Amazon.xlsx')
        self.notasCredito: List[pd.DataFrame] = [] 
        pass

    def run(self, zip_file, inputZipName):
        self.__processZip(zip_file)
        data = self.__formatOutputFiles(inputZipName)

        return data
        
    
    def __formatOutputFiles(self, inputZipName):
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as temp_factura:
            facturaFilename = temp_factura.name
            self.facturasWB.save(facturaFilename)
        
        notasCreditoFilenames: List[dict[str,str]]= []
        for notaCredito in self.notasCredito:
            with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as temp_notaCredito:
                notasCreditoFilenames.append({"temp": temp_notaCredito.name, "final": notaCredito["name"]})
                notaCredito["data"].to_excel(temp_notaCredito.name, startrow=15, index=False)

        with tempfile.NamedTemporaryFile(suffix=".zip") as temp_zip:
            with zipfile.ZipFile(temp_zip.name, 'w') as zipf:
                zipf.write(facturaFilename, arcname= inputZipName.split(".")[0] + ".xlsx")

                for i, filename in enumerate(notasCreditoFilenames):
                    zipf.write(filename["temp"], arcname= f"credit_notes/{filename['final']}.xlsx")

            zip_bytes = temp_zip.read()

        if os.path.exists(facturaFilename): # Borra temp files
            os.remove(facturaFilename)

        for notaCredito in notasCreditoFilenames:
             if os.path.exists(notaCredito["temp"]): # Borra temp files
                os.remove(notaCredito["temp"])

        st.write("# credit notes: " + str(len(notasCreditoFilenames)))
        return zip_bytes;
    
    def __processZip(self, zip_file):
        try:
            with zipfile.ZipFile(BytesIO(zip_file.read())) as zf:

                file_list = zf.namelist()
                st.write("# files: " + str(len(file_list)))

                for file_name in file_list:
                    if file_name.endswith(('.xls', '.xlsx')):
                        with zf.open(file_name) as file:
                            try:
                                self.__processFile(file, file_name)
                            except Exception as e:
                                st.error(f"Error reading {file_name}: {str(e)}")
                    else:
                        st.warning(f"Skipping non-Excel file: {file_name}")
        except zipfile.BadZipFile:
            st.error("Archivo invalido.")
        except Exception as e:
            st.error(f"Error procesando zip: {str(e)}")

    def __cleanClientName(_, client: str):
        client = client.replace(", SA DE CV", "").strip()
        client = client.replace(", S.A. DE C.V.", "").strip()
        client = client.replace(", S.A.DE C.V.", "").strip()
        client = client.replace(", S.A. DEC.V.", "").strip()
        client = client.replace(",S.A. DE C.V.", "").strip()
        client = client.replace(", S.A. DE CV.", "").strip()
        client = client.replace(" S.A. DE C.V.", "").strip()
        client = client.replace(", S.A. DE CV", "").strip()
        client = client.replace(", S.C. DE R.L. DE C.V.", "").strip()
        client = client.replace(" SA DE CV", "").strip()
        client = client.replace(",S.A.", "").strip()
        client = client.replace(",S.A", "").strip()
        client = client.replace(", S.A", "").strip()
        return client

    def __processFile(self, file, file_name):
        input_data = pd.read_excel(file, header=19)  # header is in the 20th row (0-indexed)
        input_data.columns = input_data.columns.str.strip()

        wb = openpyxl.load_workbook(file)
        sheet = wb.active

        cliente = self.__cleanClientName(sheet['A3'].value)
        rfc = sheet['A6'].value.split(": ")[-1]  # Extract the value next to ": "
        referencia = sheet['A4'].value

        output_columns = [...]  
        output_data = pd.DataFrame(columns=output_columns)
        credit_notes_data = pd.DataFrame(columns=output_columns)

        factura_number = 1
        for index, row in input_data.iterrows():
            concepto = row['Description']
            if pd.isna(concepto):
                break

            concepto = concepto.strip()
            united_price_usd = row['United Price MXN']
            codigo_del_concepto = 90111500 if "MONTHLY FEE" in concepto.upper() else 80141600
            importe = united_price_usd

            if united_price_usd < 0:
                credit_notes_data = pd.concat([credit_notes_data, pd.DataFrame([{
                    'DESPACHO': 'MIDESPACHO',
                    'NO. FACTURA': factura_number,
                    'CLIENTE': cliente,
                    'TAXID (EXTRANJERO)': None,
                    'RFC*': rfc,
                    'NOMBRE DEL CONTACTO*': None,
                    'DIRECCIÓN*': None,
                    'TELÉFONO*': None,
                    'CORREO ELECTRÓNICO*': None,
                    'CONCEPTO': concepto,
                    'CANTIDAD': 1,
                    'CÓDIGO DEL CONCEPTO': codigo_del_concepto,
                    'PARCIALIDADES': None,
                    'MONEDA': 'MXN',
                    'IMPORTE': importe,
                    'IMPUESTO': 'IVA16',
                    'REFERENCIA': referencia
                }])], ignore_index=True)
            else:
                # Add to the main output DataFrame
                output_data = pd.concat([output_data, pd.DataFrame([{
                    'DESPACHO': 'MIDESPACHO',
                    'NO. FACTURA': factura_number,
                    'CLIENTE': cliente,
                    'TAXID (EXTRANJERO)': None,
                    'RFC*': rfc,
                    'NOMBRE DEL CONTACTO*': None,
                    'DIRECCIÓN*': None,
                    'TELÉFONO*': None,
                    'CORREO ELECTRÓNICO*': None,
                    'CONCEPTO': concepto,
                    'CANTIDAD': 1,
                    'CÓDIGO DEL CONCEPTO': codigo_del_concepto,
                    'PARCIALIDADES': None,
                    'MONEDA': 'MXN',
                    'IMPORTE': importe,
                    'IMPUESTO': 'IVA16',
                    'REFERENCIA': referencia
                }])], ignore_index=True)
            factura_number += 1
        wb.close()

        output_data = output_data.drop(output_data.columns[0],axis=1)
        data_to_write = output_data.values.tolist()

        ws = self.facturasWB.active
        start_column = 3  
        for row in data_to_write:
            for idx, value in enumerate(row):
                ws.cell(row=self.start_row, column=start_column + idx, value=value)
            self.start_row += 1  # Move to the next row after writing each row of data
        self.facturasWB.close()

        if not credit_notes_data.empty:
            self.notasCredito.append({"data": credit_notes_data, "name": file_name})

        

        